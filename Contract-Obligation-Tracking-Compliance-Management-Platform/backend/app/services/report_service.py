from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from sqlalchemy.orm import Session

from app.models.contract import Contract
from app.models.obligation import Obligation
from app.models.renewal import Renewal
from app.models.audit_log import AuditLog
from app.models.user import User

from app.services.compliance_service import (
    calculate_contract_compliance,
)


# ============================================================
# Helper
# ============================================================

def _status_count(items, attribute: str, status: str) -> int:
    return sum(
        1
        for item in items
        if getattr(item, attribute, None) == status
    )


# ============================================================
# Contract Report
# ============================================================

def generate_contract_report(db: Session) -> dict[str, Any]:

    contracts = (
        db.query(Contract)
        .order_by(Contract.id.asc())
        .all()
    )

    result = {
        "total_contracts": len(contracts),
        "active_contracts": 0,
        "draft_contracts": 0,
        "under_review_contracts": 0,
        "approved_contracts": 0,
        "expired_contracts": 0,
        "terminated_contracts": 0,
        "contracts": [],
    }

    for contract in contracts:

        status = contract.status or "Draft"

        if status == "Active":
            result["active_contracts"] += 1
        elif status == "Draft":
            result["draft_contracts"] += 1
        elif status == "Under Review":
            result["under_review_contracts"] += 1
        elif status == "Approved":
            result["approved_contracts"] += 1
        elif status == "Expired":
            result["expired_contracts"] += 1
        elif status == "Terminated":
            result["terminated_contracts"] += 1

        result["contracts"].append({
            "contract_id": contract.id,
            "contract_number": contract.contract_number,
            "title": contract.title,
            "category": contract.category,
            "start_date": contract.start_date,
            "end_date": contract.end_date,
            "status": status,
        })

    return result


# ============================================================
# Obligation Report
# ============================================================

def generate_obligation_report(db: Session) -> dict[str, Any]:

    obligations = (
        db.query(Obligation)
        .order_by(Obligation.due_date.asc())
        .all()
    )

    result = {
        "total_obligations": len(obligations),
        "pending_obligations": 0,
        "in_progress_obligations": 0,
        "delayed_obligations": 0,
        "overdue_obligations": 0,
        "completed_obligations": 0,
        "obligations": [],
    }

    today = date.today()

    for obligation in obligations:

        status = obligation.status or "Pending"

        # Detect overdue obligations using due date.
        if (
            status not in {"Completed", "Delayed"}
            and obligation.due_date
            and obligation.due_date < today
        ):
            status = "Overdue"

        if status == "Pending":
            result["pending_obligations"] += 1

        elif status == "In Progress":
            result["in_progress_obligations"] += 1

        elif status == "Delayed":
            result["delayed_obligations"] += 1

        elif status == "Overdue":
            result["overdue_obligations"] += 1

        elif status == "Completed":
            result["completed_obligations"] += 1

        contract_number = None
        if obligation.contract:
            contract_number = obligation.contract.contract_number

        assigned_user = None
        if obligation.assignee:
            assigned_user = obligation.assignee.full_name

        result["obligations"].append({
            "obligation_id": obligation.id,
            "title": obligation.title,
            "obligation_type": obligation.obligation_type,
            "contract_id": obligation.contract_id,
            "contract_number": contract_number,
            "due_date": obligation.due_date,
            "assigned_to": obligation.assigned_to,
            "assigned_user": assigned_user,
            "status": status,
            "completion_date": obligation.completion_date,
        })

    return result


# ============================================================
# Renewal Report
# ============================================================

def generate_renewal_report(db: Session) -> dict[str, Any]:

    renewals = (
        db.query(Renewal)
        .order_by(Renewal.previous_expiry_date.asc())
        .all()
    )

    result = {
        "total_renewals": len(renewals),
        "upcoming_renewals": 0,
        "in_progress_renewals": 0,
        "renewed_renewals": 0,
        "expired_renewals": 0,
        "cancelled_renewals": 0,
        "renewals": [],
    }

    for renewal in renewals:

        status = renewal.status or "Upcoming"

        if status == "Upcoming":
            result["upcoming_renewals"] += 1

        elif status == "In Progress":
            result["in_progress_renewals"] += 1

        elif status == "Renewed":
            result["renewed_renewals"] += 1

        elif status == "Expired":
            result["expired_renewals"] += 1

        elif status == "Cancelled":
            result["cancelled_renewals"] += 1

        contract_number = None
        contract_title = None

        if renewal.contract:
            contract_number = renewal.contract.contract_number
            contract_title = renewal.contract.title

        assigned_user = None

        if renewal.assigned_user:
            assigned_user = renewal.assigned_user.full_name

        result["renewals"].append({
            "renewal_id": renewal.id,
            "contract_id": renewal.contract_id,
            "contract_number": contract_number,
            "contract_title": contract_title,
            "renewal_date": renewal.renewal_date,
            "previous_expiry_date": renewal.previous_expiry_date,
            "new_expiry_date": renewal.new_expiry_date,
            "status": status,
            "assigned_to": renewal.assigned_to,
            "assigned_user": assigned_user,
        })

    return result


# ============================================================
# Compliance Report
# ============================================================

def generate_compliance_report(db: Session) -> dict[str, Any]:

    contracts = db.query(Contract).all()

    result = {
        "total_contracts": len(contracts),
        "compliant_contracts": 0,
        "pending_contracts": 0,
        "delayed_contracts": 0,
        "non_compliant_contracts": 0,
        "high_risk_contracts": 0,
        "average_compliance_score": 0.0,
        "compliance_reports": [],
    }

    total_score = 0.0

    for contract in contracts:

        obligations = (
            db.query(Obligation)
            .filter(
                Obligation.contract_id == contract.id
            )
            .all()
        )

        compliance = calculate_contract_compliance(
            contract,
            obligations,
        )

        status = compliance["compliance_status"]

        if status == "Compliant":
            result["compliant_contracts"] += 1

        elif status == "Pending":
            result["pending_contracts"] += 1

        elif status == "Delayed":
            result["delayed_contracts"] += 1

        elif status == "Non-Compliant":
            result["non_compliant_contracts"] += 1

        elif status == "High Risk":
            result["high_risk_contracts"] += 1

        score = float(
            compliance.get("compliance_score", 0)
        )

        total_score += score

        result["compliance_reports"].append({
            "contract_id": compliance["contract_id"],
            "contract_number": compliance.get(
                "contract_number"
            ),
            "compliance_status": status,
            "compliance_score": score,
            "total_obligations": compliance[
                "total_obligations"
            ],
            "completed_obligations": compliance[
                "completed_obligations"
            ],
            "pending_obligations": compliance[
                "pending_obligations"
            ],
            "delayed_obligations": compliance[
                "delayed_obligations"
            ],
            "overdue_obligations": compliance[
                "overdue_obligations"
            ],
            "risk_level": compliance["risk_level"],
            "evaluated_at": compliance.get(
                "evaluated_at"
            ),
        })

    if contracts:
        result["average_compliance_score"] = round(
            total_score / len(contracts),
            2,
        )

    return result


# ============================================================
# Audit Report
# ============================================================

def generate_audit_report(db: Session) -> dict[str, Any]:

    logs = (
        db.query(AuditLog)
        .order_by(AuditLog.created_at.desc())
        .all()
    )

    result = {
        "total_audit_logs": len(logs),
        "audit_logs": [],
    }

    for log in logs:

        user_name = None

        if log.user:
            user_name = log.user.full_name

        result["audit_logs"].append({
            "audit_id": log.id,
            "user_id": log.user_id,
            "user_name": user_name,
            "action": log.action,
            "table_name": log.table_name,
            "record_id": log.record_id,
            "created_at": log.created_at,
        })

    return result


# ============================================================
# Excel Helper
# ============================================================

def create_excel_file(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> BytesIO:

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31]

    # Header
    worksheet.append(headers)

    # Data
    for row in rows:
        worksheet.append([
            "" if value is None else str(value)
            for value in row
        ])

    # Formatting
    for cell in worksheet[1]:
        cell.font = cell.font.copy(bold=True)

    for column_cells in worksheet.columns:

        max_length = 0
        column_index = column_cells[0].column

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(
                max_length,
                len(value),
            )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = min(max_length + 2, 50)

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    return output


# ============================================================
# Contract Excel
# ============================================================

def contract_excel(db: Session) -> BytesIO:

    report = generate_contract_report(db)

    rows = [
        [
            item["contract_id"],
            item["contract_number"],
            item["title"],
            item["category"],
            item["start_date"],
            item["end_date"],
            item["status"],
        ]
        for item in report["contracts"]
    ]

    return create_excel_file(
        "Contracts",
        [
            "Contract ID",
            "Contract Number",
            "Title",
            "Category",
            "Start Date",
            "End Date",
            "Status",
        ],
        rows,
    )


# ============================================================
# Obligation Excel
# ============================================================

def obligation_excel(db: Session) -> BytesIO:

    report = generate_obligation_report(db)

    rows = [
        [
            item["obligation_id"],
            item["title"],
            item["obligation_type"],
            item["contract_id"],
            item["contract_number"],
            item["due_date"],
            item["assigned_user"],
            item["status"],
            item["completion_date"],
        ]
        for item in report["obligations"]
    ]

    return create_excel_file(
        "Obligations",
        [
            "Obligation ID",
            "Title",
            "Type",
            "Contract ID",
            "Contract Number",
            "Due Date",
            "Assigned User",
            "Status",
            "Completion Date",
        ],
        rows,
    )


# ============================================================
# Renewal Excel
# ============================================================

def renewal_excel(db: Session) -> BytesIO:

    report = generate_renewal_report(db)

    rows = [
        [
            item["renewal_id"],
            item["contract_id"],
            item["contract_number"],
            item["contract_title"],
            item["renewal_date"],
            item["previous_expiry_date"],
            item["new_expiry_date"],
            item["status"],
            item["assigned_user"],
        ]
        for item in report["renewals"]
    ]

    return create_excel_file(
        "Renewals",
        [
            "Renewal ID",
            "Contract ID",
            "Contract Number",
            "Contract Title",
            "Renewal Date",
            "Previous Expiry",
            "New Expiry",
            "Status",
            "Assigned User",
        ],
        rows,
    )


# ============================================================
# Compliance Excel
# ============================================================

def compliance_excel(db: Session) -> BytesIO:

    report = generate_compliance_report(db)

    rows = [
        [
            item["contract_id"],
            item["contract_number"],
            item["compliance_status"],
            item["compliance_score"],
            item["total_obligations"],
            item["completed_obligations"],
            item["pending_obligations"],
            item["delayed_obligations"],
            item["overdue_obligations"],
            item["risk_level"],
        ]
        for item in report["compliance_reports"]
    ]

    return create_excel_file(
        "Compliance",
        [
            "Contract ID",
            "Contract Number",
            "Compliance Status",
            "Compliance Score",
            "Total Obligations",
            "Completed",
            "Pending",
            "Delayed",
            "Overdue",
            "Risk Level",
        ],
        rows,
    )


# ============================================================
# Audit Excel
# ============================================================

def audit_excel(db: Session) -> BytesIO:

    report = generate_audit_report(db)

    rows = [
        [
            item["audit_id"],
            item["user_id"],
            item["user_name"],
            item["action"],
            item["table_name"],
            item["record_id"],
            item["created_at"],
        ]
        for item in report["audit_logs"]
    ]

    return create_excel_file(
        "Audit Logs",
        [
            "Audit ID",
            "User ID",
            "User Name",
            "Action",
            "Table",
            "Record ID",
            "Created At",
        ],
        rows,
    )


# ============================================================
# PDF Helper
# ============================================================

def create_pdf_file(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> BytesIO:

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    story = []

    story.append(
        Paragraph(
            "ContractIQ",
            styles["Title"],
        )
    )

    story.append(
        Paragraph(
            title,
            styles["Heading2"],
        )
    )

    story.append(Spacer(1, 12))

    formatted_rows = [
        headers
    ]

    for row in rows:

        formatted_rows.append([
            "" if value is None else str(value)
            for value in row
        ])

    table = Table(
        formatted_rows,
        repeatRows=1,
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.grey,
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black,
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7,
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP",
            ),
        ])
    )

    story.append(table)

    document.build(story)

    output.seek(0)

    return output


# ============================================================
# PDF Reports
# ============================================================

def contract_pdf(db: Session) -> BytesIO:

    report = generate_contract_report(db)

    rows = [
        [
            item["contract_id"],
            item["contract_number"],
            item["title"],
            item["category"],
            item["start_date"],
            item["end_date"],
            item["status"],
        ]
        for item in report["contracts"]
    ]

    return create_pdf_file(
        "Contract Report",
        [
            "ID",
            "Contract No.",
            "Title",
            "Category",
            "Start",
            "End",
            "Status",
        ],
        rows,
    )


def obligation_pdf(db: Session) -> BytesIO:

    report = generate_obligation_report(db)

    rows = [
        [
            item["obligation_id"],
            item["title"],
            item["obligation_type"],
            item["contract_number"],
            item["due_date"],
            item["assigned_user"],
            item["status"],
        ]
        for item in report["obligations"]
    ]

    return create_pdf_file(
        "Obligation Report",
        [
            "ID",
            "Title",
            "Type",
            "Contract",
            "Due Date",
            "Assigned User",
            "Status",
        ],
        rows,
    )


def renewal_pdf(db: Session) -> BytesIO:

    report = generate_renewal_report(db)

    rows = [
        [
            item["renewal_id"],
            item["contract_number"],
            item["contract_title"],
            item["previous_expiry_date"],
            item["renewal_date"],
            item["new_expiry_date"],
            item["status"],
        ]
        for item in report["renewals"]
    ]

    return create_pdf_file(
        "Renewal Report",
        [
            "ID",
            "Contract",
            "Title",
            "Previous Expiry",
            "Renewal Date",
            "New Expiry",
            "Status",
        ],
        rows,
    )


def compliance_pdf(db: Session) -> BytesIO:

    report = generate_compliance_report(db)

    rows = [
        [
            item["contract_id"],
            item["contract_number"],
            item["compliance_status"],
            item["compliance_score"],
            item["total_obligations"],
            item["completed_obligations"],
            item["overdue_obligations"],
            item["risk_level"],
        ]
        for item in report["compliance_reports"]
    ]

    return create_pdf_file(
        "Compliance Report",
        [
            "Contract ID",
            "Contract",
            "Status",
            "Score",
            "Total",
            "Completed",
            "Overdue",
            "Risk",
        ],
        rows,
    )


def audit_pdf(db: Session) -> BytesIO:

    report = generate_audit_report(db)

    rows = [
        [
            item["audit_id"],
            item["user_name"],
            item["action"],
            item["table_name"],
            item["record_id"],
            item["created_at"],
        ]
        for item in report["audit_logs"]
    ]

    return create_pdf_file(
        "Audit Report",
        [
            "ID",
            "User",
            "Action",
            "Table",
            "Record ID",
            "Created At",
        ],
        rows,
    )